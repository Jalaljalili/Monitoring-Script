import asyncio
import ipaddress
import time
import os
import subprocess
from collections import defaultdict, deque
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ------------------ Config ------------------
IP_FILE = "ip.txt"
REPORT_FILE = "port_scan_report.xlsx"

# Timeouts / concurrency
CONNECT_TIMEOUT = 3.0
MAX_CONCURRENT_HOSTS = 8            # total hosts scanned in parallel
PER_HOST_WORKERS = 200              # concurrent port dials per host
GLOBAL_MAX_CONNECTIONS = 1200       # absolute global connection cap

# Port range
PORT_START = 1
PORT_END = 65535

# Record closed/filtered rows in Excel?
REPORT_OPEN_PORTS_ONLY = True

# TLS ports to attempt direct TLS handshake (no STARTTLS here)
TLS_CANDIDATE_PORTS = {443, 8443, 993, 995, 465, 990}

# OpenSSL per-version flags
TLS_VERSIONS = {
    "SSLv3":  "-ssl3",
    "TLSv1.0": "-tls1",
    "TLSv1.1": "-tls1_1",
    "TLSv1.2": "-tls1_2",
    "TLSv1.3": "-tls1_3",
}

# Show some diagnostic warnings (too many files, etc.)
VERBOSE_WARNINGS = True
# --------------------------------------------

# Some OS errno codes for "connection refused"
REFUSED_ERRNOS = {111, 61, 10061}     # Linux=111, macOS=61, Windows=10061
TOO_MANY_FILES_ERRNO = 24

global_conn_semaphore = None  # set in main()


def check_tls_versions(ip: str, port: int) -> List[str]:
    """Try specific TLS versions with openssl s_client. Return supported versions."""
    supported = []
    for name, flag in TLS_VERSIONS.items():
        try:
            # -servername helps SNI servers; harmless for IPs
            cmd = ["openssl", "s_client", "-connect", f"{ip}:{port}", flag, "-servername", ip, "-brief"]
            proc = subprocess.run(cmd, capture_output=True, text=True, timeout=6)
            out = (proc.stdout or "") + "\n" + (proc.stderr or "")
            # Heuristic: success if it didn’t say failure and we see signs of a session
            failed_markers = ["handshake failure", "wrong version number", "no protocols available", "alert"]
            if any(m in out.lower() for m in failed_markers):
                continue
            if "Protocol" in out or "SSL-Session:" in out or "Server certificate" in out or "Cipher is" in out:
                supported.append(name)
        except subprocess.TimeoutExpired:
            continue
        except Exception:
            continue
    return supported


async def dial_port(host: str, port: int, results_queue: asyncio.Queue):
    """TCP connect scan for a single port with global concurrency cap."""
    try:
        async with global_conn_semaphore:
            reader = writer = None
            try:
                reader, writer = await asyncio.wait_for(
                    asyncio.open_connection(host, port),
                    timeout=CONNECT_TIMEOUT
                )
                # Connected -> OPEN
                await results_queue.put((host, port, "OPEN"))
            except OSError as e:
                # Host answered with RST -> REFUSED (host is up!)
                if getattr(e, "errno", None) in REFUSED_ERRNOS:
                    await results_queue.put((host, port, "REFUSED"))
                else:
                    if VERBOSE_WARNINGS and getattr(e, "errno", None) == TOO_MANY_FILES_ERRNO:
                        print("[WARN] Too many open files; lower concurrency or raise ulimit -n")
                    if not REPORT_OPEN_PORTS_ONLY:
                        await results_queue.put((host, port, f"ERROR:{getattr(e, 'errno', None)}"))
            except asyncio.TimeoutError:
                if not REPORT_OPEN_PORTS_ONLY:
                    await results_queue.put((host, port, "TIMEOUT"))
            finally:
                if writer is not None:
                    writer.close()
                    try:
                        await writer.wait_closed()
                    except Exception:
                        pass
    except Exception:
        if not REPORT_OPEN_PORTS_ONLY:
            await results_queue.put((host, port, "ERROR"))


async def scan_host(host: str, results_queue: asyncio.Queue):
    """Scan all ports on a host using a bounded worker pool."""
    print(f"[*] Scanning host: {host}")

    port_queue: asyncio.Queue = asyncio.Queue()
    for p in range(PORT_START, PORT_END + 1):
        port_queue.put_nowait(p)

    async def worker():
        while True:
            try:
                port = port_queue.get_nowait()
            except asyncio.QueueEmpty:
                return
            await dial_port(host, port, results_queue)
            port_queue.task_done()

    workers = [asyncio.create_task(worker()) for _ in range(PER_HOST_WORKERS)]
    await port_queue.join()
    for w in workers:
        await w

    print(f"[*] Finished scanning host: {host}")


async def main():
    global global_conn_semaphore
    global_conn_semaphore = asyncio.Semaphore(GLOBAL_MAX_CONNECTIONS)

    started = time.time()

    # ---- Load IPs ----
    if not os.path.exists(IP_FILE):
        print(f"[!] ip file not found: {IP_FILE}")
        return

    all_ips = []
    with open(IP_FILE, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            try:
                net = ipaddress.ip_network(line, strict=False)
                for ip in net.hosts():
                    all_ips.append(str(ip))
            except ValueError:
                # Single IP not in CIDR form? Accept as host.
                try:
                    ipaddress.ip_address(line)
                    all_ips.append(line)
                except ValueError:
                    print(f"[!] Invalid entry in {IP_FILE}: {line}")

    if not all_ips:
        print("[!] No valid IPs to scan.")
        return

    print(f"[+] Total IPs to scan: {len(all_ips)}")
    results_queue: asyncio.Queue = asyncio.Queue()

    # Shared results structure
    scan_results: Dict[str, Dict] = defaultdict(lambda: {
        "status": "Down",
        "open_ports": set(),
        "tls_by_port": {}  # port -> [versions]
    })

    # ---- Schedule hosts with a cap on concurrent hosts ----
    hosts = deque(all_ips)
    active = set()

    while hosts or active:
        while hosts and len(active) < MAX_CONCURRENT_HOSTS:
            h = hosts.popleft()
            task = asyncio.create_task(scan_host(h, results_queue))
            task.host = h  # tag for logging
            active.add(task)
            # ensure appears in output even if nothing found
            _ = scan_results[h]

        if active:
            done, pending = await asyncio.wait(active, timeout=0.1, return_when=asyncio.FIRST_COMPLETED)
            for t in done:
                active.remove(t)
                try:
                    await t
                except Exception as e:
                    print(f"[!] Host task error for {getattr(t, 'host', '?')}: {e}")

        # Drain results frequently
        while not results_queue.empty():
            host, port, status = await results_queue.get()
            if status == "OPEN":
                scan_results[host]["status"] = "Up"
                scan_results[host]["open_ports"].add(port)
                # TLS check for candidate ports
                if port in TLS_CANDIDATE_PORTS and port not in scan_results[host]["tls_by_port"]:
                    scan_results[host]["tls_by_port"][port] = check_tls_versions(host, port)
            elif status == "REFUSED":
                # Even if no open ports, a refused connection means host is up
                if scan_results[host]["status"] != "Up":
                    scan_results[host]["status"] = "Up"
            else:
                # CLOSED/FILTERED/TIMEOUT/ERROR (only if REPORT_OPEN_PORTS_ONLY=False)
                pass
            results_queue.task_done()

        await asyncio.sleep(0.002)

    # Final drain
    await results_queue.join()
    while not results_queue.empty():
        host, port, status = await results_queue.get()
        if status == "OPEN":
            scan_results[host]["status"] = "Up"
            scan_results[host]["open_ports"].add(port)
        elif status == "REFUSED":
            if scan_results[host]["status"] != "Up":
                scan_results[host]["status"] = "Up"
        results_queue.task_done()

    elapsed = time.time() - started
    print(f"[+] Scan finished in {elapsed:.2f}s")

    # ------------- Excel report -------------
    print(f"[+] Writing Excel: {REPORT_FILE}")
    wb = Workbook()
    sh = wb.active
    sh.title = "Port Scan Results"

    headers = ["IP Address", "Host Status", "Open Ports", "TLS Versions (by port)"]
    sh.append(headers)
    for c in sh[1]:
        c.font = Font(bold=True)

    def tls_cell_str(tls_by_port: Dict[int, List[str]]) -> str:
        if not tls_by_port:
            return "N/A"
        parts = []
        for p in sorted(tls_by_port):
            vers = tls_by_port[p] or ["None"]
            parts.append(f"{p}: {', '.join(vers)}")
        return " | ".join(parts)

    for ip in sorted(scan_results.keys(), key=lambda s: ipaddress.IPv4Address(s)):
        data = scan_results[ip]
        open_ports = ", ".join(map(str, sorted(data["open_ports"]))) if data["open_ports"] else "None"
        tls_str = tls_cell_str(data["tls_by_port"])

        row = [ip, data["status"], open_ports, tls_str]
        sh.append(row)
        status_cell = sh.cell(row=sh.max_row, column=2)
        if data["status"] == "Up":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Auto width
    for col in sh.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        sh.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(REPORT_FILE)
    print(f"[+] Report saved: {REPORT_FILE}")
    # ---------------------------------------
if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n[!] Interrupted.")
